# -*- coding: utf-8 -*-
"""
Парсер Excel с учётом нескольких пустых строк-разделителей.
"""

import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def _cell_value(v: Any) -> str:
    """Нормализуем значение ячейки в строку."""
    if v is None:
        return ""
    return str(v).strip()


def _row_is_empty(ws: Worksheet, row_idx: int, max_cols: int = 8) -> bool:
    """Проверяем, что строка пустая по первым max_cols колонкам."""
    for col in range(1, max_cols + 1):
        if _cell_value(ws.cell(row=row_idx, column=col).value) != "":
            return False
    return True


def _join_8(row_vals: List[str]) -> str:
    """Собираем 8 значений в строку с кавычками и \t."""
    safe = [v.replace('"', r'\"') for v in row_vals]
    quoted = [f"\"{v}\"" for v in safe]
    return "\t".join(quoted)

MAX_COLS = 10

def parse_excel_to_json(
    xlsx_path: Union[str, Path],
    sheet_name: Optional[str] = None
) -> list:
    wb = load_workbook(filename=xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    results: List[Dict[str, Any]] = []

    r = 1
    max_row = ws.max_row

    while r <= max_row:
        first_col_val = _cell_value(ws.cell(row=r, column=1).value)

        # Ищем заголовок прибора
        if first_col_val:
            device_name = first_col_val

            # Описание — строка ниже
            desc_row = r + 1
            device_desc = _cell_value(ws.cell(row=desc_row, column=1).value) if desc_row <= max_row else ""

            # Пропускаем 2 строки шапки
            start_table_row = r + 3
            table_rows: List[str] = []

            cur = start_table_row
            prev_vals: List[str] = [""] * MAX_COLS  # храним последние значения по колонкам

            while cur <= max_row:
                if _row_is_empty(ws, cur, max_cols=MAX_COLS):
                    break

                vals = []
                for c in range(1, MAX_COLS + 1):
                    cell_val = _cell_value(ws.cell(row=cur, column=c).value)
                    if cell_val == "" and prev_vals[c - 1] != "":
                        # берём из предыдущей строки
                        vals.append(prev_vals[c - 1])
                    else:
                        vals.append(cell_val)

                prev_vals = vals[:]  # обновляем "последние значения"
                line = _join_8(vals)
                table_rows.append(line)

                cur += 1

            results.append(
                {
                    "device": device_name,
                    "description": device_desc,
                    "rows": table_rows,
                }
            )

            # === Пропускаем все пустые строки-разделители ===
            while cur <= max_row and _row_is_empty(ws, cur, max_cols=MAX_COLS):
                cur += 1

            r = cur
            continue

        # Если первая колонка пустая — просто идём дальше
        r += 1

    return results


import re

def convert(data: list) -> list:
    def parse_images(s: str) -> list[str]:
        """
        Извлекает имена файлов изображений из строки.
        Поддерживает расширения jpg|jpeg|png|gif.
        """
        return re.findall(r'\b[\w-]+\.(?:jpg|jpeg|png|gif)\b', s, flags=re.IGNORECASE)
    def parse_row(text: str) -> dict:
        parts = [p.strip().strip('"') for p in text.split("\t") if p.strip()]

        # Проверяем количество частей
        required_parts = MAX_COLS  # у нас минимум до parts[9]
        if len(parts) < required_parts:
            raise ValueError(
                f"Ошибка парсинга строки: ожидалось минимум {required_parts} частей, "
                f"но получено {len(parts)}.\nСтрока: {text}"
            )

        row = {}

        # 1. Модель
        model_raw = parts[0].strip()
        if "\n" in model_raw:
            lines = [line.strip() for line in model_raw.split("\n") if line.strip()]
            first, rest = lines[0], lines[1:]
            formatted = [first]

            for r in rest:
                # считаем количество заглавных букв
                upper_count = sum(1 for ch in r if ch.isupper())
                if upper_count >= 2:
                    formatted.append(f"<br>{r}")
                else:
                    formatted.append(f"<br><span class='table_subtext'>{r}</span>")

            row["Модель"] = "".join(formatted)
        else:
            row["Модель"] = model_raw

        # 2. Диаметр
        diameter_match = re.search(r"d\.?(\d+)", parts[1])
        if diameter_match:
            row["Диаметр"] = f"{diameter_match.group(1)} мм"

        # 3. Класс точности
        kt_raw = re.sub(r"к\.т\.\s*", "", parts[2])
        kt_values = [v.strip().replace("¹", "") for v in re.split(r"[;]", kt_raw) if v.strip()]
        kt_formatted = []
        for i, val in enumerate(kt_values):
            if i == 0:
                kt_formatted.append(val)
            else:
                kt_formatted.append(f"<span class='optional'>{val}¹</span>")
        row["Класс точности"] = "<br>".join(kt_formatted)

        # 4. Степень IP
        ip_values = re.findall(r"IP\d+", parts[3])
        ip_formatted = []
        for i, val in enumerate(ip_values):
            if i == 0:
                ip_formatted.append(val)
            else:
                ip_formatted.append(f"<span class='optional'>{val}¹</span>")
        row["Степень IP"] = "<br>".join(ip_formatted)

        # 5. Резьба
        thread_values = re.split(r"[; ]+", parts[4].replace("¹", ""))
        thread_values = [v for v in thread_values if v]
        thread_formatted = []
        for i, val in enumerate(thread_values):
            if i == 0:
                thread_formatted.append(val)
            else:
                thread_formatted.append(f"<span class='optional'>{val}¹</span>")
        row["Резьба"] = "<br>".join(thread_formatted)

        # 6. Климат
        climate_raw = parts[5]
        base, *opts = re.split(r"[()]", climate_raw)
        climate_values = [base.strip()]
        if opts:
            for opt in opts[0].split(";"):
                opt = opt.strip().replace("¹", "")
                if opt:
                    climate_values.append(f"<span class='optional'>{opt}¹</span>")
        row["Климат"] = "<br>".join(climate_values)

        # 7. Виброзащита
        vibro_raw = parts[6]
        if re.search(r"[А-Яа-яЁё]", vibro_raw):
            # ищем содержимое в скобках и заменяем с ¹
            def replacer(match):
                opt = match.group(1)
                suffix = "¹" if match.group(2) else ""
                return f"<span class='optional'>({opt}){suffix}</span>"

            # ( ... ) + опционально '1' или '¹'
            result = re.sub(r"\(([^)]*?)\)\s*([1¹])?", replacer, vibro_raw)

            row_value = result.strip()
        else:
            # старое поведение
            base, *opts = re.split(r"[()]", vibro_raw)
            vibro_values = [base.strip()]
            if opts:
                for opt in opts[0].split(";"):
                    opt = opt.strip().replace("¹", "")
                    if opt:
                        vibro_values.append(f"<span class='optional'>{opt}¹</span>")
            row_value = "<br>".join(vibro_values)

        row["Вибро защита"] = row_value

        # 8. Пределы давления
        pressure_raw = parts[7]
        pressure_lines = [line.strip() for line in pressure_raw.split("\n") if line.strip()]
        result = []
        for line in pressure_lines:
            chunks = re.split(r'([A-Za-zА-Яа-яЁё]+;)', line)
            buf = ""
            for chunk in chunks:
                buf += chunk
                if re.fullmatch(r'[A-Za-zА-Яа-яЁё]+;', chunk):
                    result.append(buf.strip())
                    buf = ""
            if buf.strip():
                result.append(buf.strip())
        row["Пределы давления"] = result

        # 9. Диапазон температур (сохраняем отдельно, чтобы вынести на уровень params)
        row["_temp_measured"] = parts[8].strip().replace('\n', ', ') + '.'
        row["_temp_env"] = parts[9].strip().replace('\n', ', ') + '.'

        return row

    result = []
    for group in data:
        if "по заказу" in group["device"]:
            continue

        parsed_rows = []
        images = []

        for row in group["rows"]:
            if "зображен" in row:
                images.extend(parse_images(row))
            else:
                parsed_rows.append(parse_row(row))

        # Берём диапазоны температур из первой строки (или можно объединять все, если нужно)
        if parsed_rows:
            temp_measured = parsed_rows[0].pop("_temp_measured", None)
            temp_env = parsed_rows[0].pop("_temp_env", None)
        else:
            temp_measured = temp_env = None

        new_group = {
            "device": group["device"],
            "description": group["description"],
            "images": images,
            "params": {
                "Диапазон температур измеряемой среды": temp_measured,
                "Диапазон температур окружающей среды": temp_env,
            },
            "data": parsed_rows
        }
        result.append(new_group)

    return result

def process(in_xlsx: str, out_json: str):
    sheet = None
    data = convert(parse_excel_to_json(in_xlsx, sheet))
    Path(out_json).write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Сохранено: {out_json}. Найдено блоков: {len(data)}")
